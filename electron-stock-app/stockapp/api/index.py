import asyncio
from pyppeteer import launch
import requests
from bs4 import BeautifulSoup
import time
from openpyxl import load_workbook
import sys

global url
global keyword
global price
StockName = ""
price = ""
Sname = ""


def work(filename):
    global price
    global total_rows
    global total_cols
    global Sname
    global Scriptoutput

    try:
        wb = load_workbook(filename)
        worksheet = wb.active

        total_rows = worksheet.max_row
        total_cols = worksheet.max_column

        for i in range(2, total_rows+1):
            Sname = worksheet.cell(row=i, column=1)
            Sname = str(Sname.value)
            if Sname != None:
                keyword = Sname + ' share price'
                # print(keyword)
                # google_share(Sname)

                soup = asyncio.get_event_loop().run_until_complete(main(keyword))
                price, open, high, low, mkt_cap, p_e_ratio, Div_yield, pre_close, wk_high, wk_low, status, reason, stock = get_data(
                    soup)
                #price = get_data(soup)
                cell_obj = worksheet.cell(row=i, column=2)
                cell_obj.value = price
                cell_obj = worksheet.cell(row=i, column=2)
                cell_obj.value = stock
                cell_obj = worksheet.cell(row=i, column=4)
                cell_obj.value = open
                cell_obj = worksheet.cell(row=i, column=5)
                cell_obj.value = high
                cell_obj = worksheet.cell(row=i, column=6)
                cell_obj.value = low
                cell_obj = worksheet.cell(row=i, column=7)
                cell_obj.value = mkt_cap
                cell_obj = worksheet.cell(row=i, column=8)
                cell_obj.value = pre_close
                cell_obj = worksheet.cell(row=i, column=9)
                cell_obj.value = status
                cell_obj = worksheet.cell(row=i, column=10)
                cell_obj.value = reason
                i += 1

                wb.save(filename)
    except Exception as e:
        pass
        print(e)


async def main(keyword):

    browser = await launch()
    page = await browser.newPage()
    await page.goto("https://www.google.co.in/search?q={}".format(keyword))
    data = await page.content()
    soup = BeautifulSoup(data, 'html.parser')
    # with open('pydata.html', 'wb') as t:
    #     t.write(soup.prettify().encode('utf-8'))
    await browser.close()
    return soup


def get_data(soup):
    try:
        price = soup.find(
            'span', class_='knowledge-finance-wholepage-chart__hover-card-value').text
        # print('price--->',price)
        data = soup.findAll('div', class_='ZSM8k')
        stock = soup.find('span', class_='HfMth').text
        # print('stock--->',stock)
        data1 = data[0]
        first_col_data = data1.findAll('tr')
        open = first_col_data[0].findAll('td')[1].text
        high = first_col_data[1].findAll('td')[1].text
        low = first_col_data[2].findAll('td')[1].text
        mkt_cap = first_col_data[3].findAll('td')[1].text
        p_e_ratio = first_col_data[4].findAll('td')[1].text

        data2 = data[1]
        second_col_data = data2.findAll('tr')
        Div_yield = second_col_data[0].findAll('td')[1].text
        pre_close = second_col_data[1].findAll('td')[1].text
        wk_high = second_col_data[2].findAll('td')[1].text
        wk_low = second_col_data[3].findAll('td')[1].text

        # print('open---->',open)
        # print('high---->',high)
        # print('low---->',low)
        # print('mkt_cap---->',mkt_cap)
        # print('p_e_ratio---->',p_e_ratio)
        # print('Div_yield---->',Div_yield)
        # print('pre_close---->',pre_close)
        # print('wk_high---->',wk_high)
        # print('wk_low---->',wk_low)
        status = 'Succesful'
        reason = 'Data found'
        return price, open, high, low, mkt_cap, p_e_ratio, Div_yield, pre_close, wk_high, wk_low, status, reason, stock
    except:
        price = None
        open = None
        high = None
        low = None
        mkt_cap = None
        p_e_ratio = None
        Div_yield = None
        pre_close = None
        wk_high = None
        wk_low = None
        status = 'Failed'
        reason = 'Data not found'
        stock = ''
        return price, open, high, low, mkt_cap, p_e_ratio, Div_yield, pre_close, wk_high, wk_low, status, reason, stock


if __name__ == '__main__':
    if sys.argv[1] == 'path':
        work(sys.argv[2])
        print('successful')
    elif sys.argv[1] == 'keyword':
        print(sys.argv[2])
        soup = asyncio.get_event_loop().run_until_complete(
            main(sys.argv[2]+"share price"))
        price, open, high, low, mkt_cap, p_e_ratio, Div_yield, pre_close, wk_high, wk_low, status, reason, stockock = get_data(
            soup)
        print('Price:', price)
        print('Open:', open)
        print('High:', high)
        print('Low:', low)
        print('mkt  cap:', mkt_cap)
        print('p.e  ratio:', p_e_ratio)
        print('Div  yield:', Div_yield)
        print('pre  close:', pre_close)
        print('wk  high:', wk_high)
        print('wk  low:', wk_low)
    else:
        print('someting went wrong')
